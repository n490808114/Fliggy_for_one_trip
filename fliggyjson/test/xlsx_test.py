from openpyxl import load_workbook
# 打开xlsx
wb = load_workbook('test.xlsx')
# 打印所有工作表的名称
print(wb.sheetnames)
# 获取活跃工作表
ws = wb.active
# 以工作表名称获取工作表
ws = wb['result']
# 更改工作表名称
ws.title = 'num'



# 保存xlsx
wb.save('test.xlsx')
